#!/usr/bin/env python
# coding: utf-8

# # Acceleration Data Processor (Sets)
# 
# ## How to use
# 1. Make sure all the required packages are installed
#     * Run the following command to install if you haven't: ```pip install pandas matplotlib scipy openpyxl notebook```
# 2. If you are seeing this, you are already running jupyter notebook
# 3. To run this program, navigate to:
#     1. Menu Bar >
#     2. Kernel >
#     3. Restart & Run All
# 4. Follow the prompts at the bottom of this notebook

# # 1) Import modules

# In[ ]:


import pandas as pd
pd.options.mode.chained_assignment = None 

import math
import csv
import os.path
import glob
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from scipy.signal import lfilter

from aux_fn import *


# # 2) Data Conversions

# ## Converts CSV to Excel Workbook Format

# In[ ]:


def csvToExcel(d, output_path, is_ms):
    '''
    Input: A dictionary of .csv file names, title of file name
    Output: Data imported into excel workbook
    '''
    
    # Create new Excel workbook
    wb = openpyxl.Workbook()
    # Just to remove the spare sheet
    wb.remove(wb.active)
    # Get the column ids where the required data is located from the user
    startIndex, interval, column_ids = getInputForColumns()
    excel_wb_name = "working-merged"

    # Transform the user input in A1 format into index for use later
    # [0: Time-stamp, 1: Acceleration X, 2: Acceleration Y, 3: Acceleration Z]
    column_indexes = [a1ToIndex(colID, True) for colID in column_ids]

    
    # For each key in dictionary
    for i, key in enumerate(d):
        # Concatenate csv files which start and end at same point together
        df_list = []
        for file in d[key]:
            formatted_output_path = output_path + "/" + file + "-formatted.csv"

            # Creates a clean csv file for output
            with open(formatted_output_path,"w",newline='') as f:
                with open(f'{file}.csv') as my_file:
            
                    # Sets this as csv
                    reader = csv.reader(my_file)
                    formatted_csv = csv.writer(f)

                    # Create a header in formatted.
                    header = ['TimeStamp','TimeElapsed','Category','Acc-X','Acc-Y','Acc-Z']
                    formatted_csv.writerow(header)

                    # Start filtering
                    i = 0 #initial value for the index, used for mapping
                    for item in reader:                        
                        if ((i-startIndex) >= 0) & (((i - startIndex) % interval) == 0):
                            each_line = []
                            for column_index in column_indexes:
                                try:                                    
                                    each_line.append(item[column_index])
                                except:
                                    pass # To prevent the code from stopping if some cells are blank
                            each_line.insert(1,0)
                            each_line.insert(2,'Acceleration')
                            formatted_csv.writerow(each_line)
                            i += 1
                        else:
                            i += 1
                formatted_df = pd.read_csv(formatted_output_path,header = 0)
                timeElapsed(formatted_df, is_ms)
                #     Convert the respective files into a list of df
                df_list.append(formatted_df)


    #     Concatenate the dfs into 1
        concat_df = pd.concat(df_list, axis = 1)

    #     Create a new column: Average of all acceleration data
        try:
            concat_df['Average Y'] = concat_df['Acc-Y'].mean(axis=1)
        except:
    #         If only have 1 set of readings, still need to create a new column
            concat_df['Average Y'] = concat_df['Acc-Y']
      
    #     Store key as title of file
        title = key

    #     Create ws with the title == key of dictionary
        ws = wb.create_sheet(index=i, title=title)

    #     Import the df into the excel sheet
        for r in dataframe_to_rows(concat_df, index=False, header=True):
            ws.append(r)

    excel_wb_name = f"{output_path}/{excel_wb_name}"
    wb.save(f'{excel_wb_name}.xlsx')
    print(f"Excel workbook created at {output_path}")
    return excel_wb_name


# # 3) DataFrame Manipulation

# ## Function to group csv files with the same start and destination together

# In[ ]:


def hasNumbers(inputString):
    return any(char.isdigit() for char in inputString)


# In[ ]:


def sortedDict(all_files):
    '''
    Input: Requires a list of .csv files
    Output:
    - A dictionary with values sorted according to the train trip
    - file_names excluding '.csv'
    
    '''

    d = {}

    # Retrieve file_names
    file_names = ''.join(all_files).split('.csv')[:-1]


    for name in file_names:

    # If name of file has a digit, find index of digit
        if hasNumbers(name):

    # Find the index of the digit
            idx = [char.isdigit() for char in name].index(True)

    # Split by the index and store the front end as the key
    # Note: '-1' is to exclude the whitespace as well
            key = name[:idx-1]
            if key in d:
                d[key].append(name)
            else:
                d[key] = [name]
        else:
            if name in d:
                d[name].append(name)
            else:
                d[name] = [name]

    
    return d, file_names


# ## Function to remove data for when train is stationary

# In[ ]:


def truncateStationaryValues(df, Threshold, isInG):
    # Currently using a rolling average n of 5.
    
    threshold = Threshold # Threshold is assigned here. Typically 0.02 if values are in m/s^2
    truncate_index = 0
    rolling_ave = df["Average Y"].rolling(window=5).mean().dropna()
    if isInG:
        rolling_ave *= -9.81
    
    for i in range(0, len(rolling_ave)):
        if (i+1) == len(rolling_ave):
            return "Could not find a difference that is more than threshold. Try setting a smaller threshold value."
        delta = abs(rolling_ave.iloc[i+1]) - abs(rolling_ave.iloc[i])
        if delta > threshold:
            truncate_index = i
            break
    
    rowsToTruncate = [idx for idx in range(1, truncate_index)]
    df = df.drop(rowsToTruncate)
    return df


# ## Function to Generate Time Elapsed

# In[ ]:


def timeElapsed(df, is_ms): 
    
    # Calculate and write a time_elapsed column
    df["TimeElapsed"] = df["TimeStamp"] - df["TimeStamp"].iloc[0]
    
    if is_ms == "y":
        df["TimeElapsed"] = df["TimeElapsed"] / 1000


# ## Function to Generate Time Interval

# In[ ]:


def retrieveLongestTimeCol(df, col):
    '''
    To retrieve the col inputted with the MOST number of rows
    '''
    
# Remove all rows where all values are blank then Retrieve the last value    
    last_value_of_t = df[col].dropna(how="all").iloc[-1]
    t = df[col].dropna(how="all").reset_index(drop=True) # Shift the index back to 0!

# If there's only 1 "Timestamp" in the excel sheet, then our x-axis will just equal to it. 'pass' because values assigned to t already is correct
    if type(last_value_of_t) == float or type(last_value_of_t) == int:
        pass
    else:
# Retrieve index of time that is not 'NaN'. This is equivalent to retrieving maximum time (I.e. Most number of rows)
        time_formatted = [i for i, el in enumerate(list(t.iloc[-1])) if not math.isnan(el)]
        idx = time_formatted[0] # This works because there should only be 1 value that is 'NaN' 

# t = Col with most number of rows
        t = t.iloc[:, idx]
    
    return t


# In[ ]:


def timeInterval(df):
    '''calculate time interval from the TimeElpased column'''
    t = retrieveLongestTimeCol(df, col = "TimeElapsed")
    
    # Retrieve all rows excluding last & reset the index (Shift it back down to 0)
    next_value = t.iloc[1:].reset_index(drop=True)
    # Find difference
    timeInterval = next_value - t
    # Add '0' value to first row & Format the index again
    timeInterval = pd.concat([pd.Series([0]), timeInterval]).reset_index(drop = True)
    # Apparently, df["TimeInterval"] values start from 1, hence the need for the code below
    timeInterval.index = range(1,len(timeInterval)+1)
    
    df["TimeInterval"] = timeInterval

    


# ## Function to generate Acc-Y-Adjusted

# In[ ]:


def accAdjust(df, smoothen_curve, format_acc):
    '''Generate the Acc-Y-Adjusted and take the shift value into account'''

    if format_acc == 'y':
        df["Average Y"] = df["Average Y"] * -9.81
    acc_Y = df["Average Y"]
        
    # Filter to remove noise
    n = int(smoothen_curve)  # the larger n is, the smoother the curve will be
    b = [1.0 / n] * n # b, numerator coefficient vector in a 1-D sequence
    a = 1 # a, denominator coefficient vector in a 1-D sequence
    acc_Y_filtered = lfilter(b,a,acc_Y)
        
    df["Acc-Y-Adjusted"] = acc_Y_filtered
    
    #shift value
    shift_value = df["Acc-Y-Adjusted"].mean()
    df["Acc-Y-Adjusted"] = df["Acc-Y-Adjusted"] - shift_value


# ## Function to generate V-btw2

# In[ ]:


def dv(df):
    '''Generate the d_v using the definition of integration using the Acc-Y-Adjusted values'''
    
    d_v = 0.5 * (df["Acc-Y-Adjusted"].iloc[1:].reset_index(drop=True) + df["Acc-Y-Adjusted"]) * df["TimeInterval"]
    
    df["V-btw2"] = d_v


# ## Function to generate V(t)

# In[ ]:


def vt(df):
    '''Generate the v(t) column to record the instantaneous velocity of all data points'''
    
    false_list = [0]*len(df)
    df["V(t)"] = false_list
    ins_v = [0]
    for i in range(1,len(df)):
    # index8: V-btw2
        ins_v.append(sum(df["V-btw2"].iloc[:i-1]))
    df["V(t)"] = ins_v


# ## Function to generate S-btw2

# In[ ]:


def ds(df):
    '''Generate the d_s using the definition of integration using the V(t) values'''
    
    d_s = 0.5 * (df["V(t)"].iloc[1:].reset_index(drop=True) + df["V(t)"]) * df["TimeInterval"]
    
    df["S-btw2"] = d_s


# ## Function to generate S(t)

# In[ ]:


def st(df):
    '''Generate the s(t) column to record the instantaneous displacement of all data points'''
    
    false_list = [0]*len(df)
    df["S(t)"] = false_list  
    
    ins_s = [0]
    for i in range(1,len(df)):
        ins_s.append(sum(df["S-btw2"].iloc[:i-1]))

    df["S(t)"] = ins_s


# ## Function to integrate all the functions in this section

# In[ ]:


def formatHeaders(df):
    
# Format headers as long as the first row is not a numerical value
# Note: Headers are not considered as the first row
# Note: '.item()' is to convert numpy types to native Python types
    try:
        if type(df.iloc[0, 0].item()) == float:
            pass
    except:
        header = df.iloc[0]
        df = df[1:]
        df.columns = header
    return df


# In[ ]:


def dfFormat(df, smoothen_curve, format_acc, dfFormatArgs):
    df = formatHeaders(df)
    
    timeInterval(df)
    
    df = truncateStationaryValues(df, dfFormatArgs["Threshold"], dfFormatArgs["isInG"])
    df = df.reset_index(drop=True)
    
    accAdjust(df, smoothen_curve, format_acc)
    dv(df)
    vt(df)
    ds(df)
    st(df)
#     print(df)
    return df


# # 4) Physical Property Determination (Data Insights)

# ## Total Distance Travelled

# In[ ]:


def distance(df):
    '''Get the total distance travelled'''
    
    Total_Distance_Travelled = max(df["S(t)"])
    
    return Total_Distance_Travelled


# ## Total Time Taken

# In[ ]:


def totalTime(df):
    '''Get the total time taken'''
    t = retrieveLongestTimeCol(df, col = "TimeElapsed")
    Total_Time_Taken = t.iloc[-1]
    
    return Total_Time_Taken


# ## Max Velocity

# In[ ]:


def maxVelocity(df):
    '''Get the max velocity'''
    
    Max_Velocity = max(df["V(t)"])
    
    return Max_Velocity


# ## Acceleration Cut-off Time & Average Acceleration

# Variables: <br>
# negative_acc_list: A list of indexes for which the acceleration value is nagetive

# In[ ]:


def aveAcc(df):
    '''Get the cut-off time for acceleration & Average Acceleration'''
    
    # To avoid early cut-off set the threshhold to be 5% of total journey
    cutoff_threshold = 0.05
    
    # A list of indexes for which the acceleration value is nagetive
    negative_acc_list = df[df["Acc-Y-Adjusted"] < 0].index.tolist()

    while negative_acc_list[0] < cutoff_threshold*len(df):
        negative_acc_list = negative_acc_list[1:]
        
    CutOff_Time = df.iloc[negative_acc_list[0],1]
    
    cutoff_idx = negative_acc_list[0]
    acc_list = []
    for i in range(cutoff_idx):
        acc_list.append(df["Acc-Y-Adjusted"].iloc[i])
    Average_Acceleration = sum(acc_list)/len(acc_list)
    
    return CutOff_Time, Average_Acceleration


# ## Max Acceleration

# In[ ]:


def maxAcc(df):
    '''Get the Maximum Acceleration'''
    
    Max_Acceleration = max(df["Acc-Y-Adjusted"])

    return Max_Acceleration


# ## Function to apply all the functions in this section

# In[ ]:


def getDataInsights(df):
    '''
    Used to get all data insights in one function
    '''
    
    totalDistTravelled = distance(df)
    totalTimeTaken = totalTime(df)
    maxV = maxVelocity(df)
    cutOffTime, aveA = aveAcc(df)
    maxA = maxAcc(df)
    data = {"d": totalDistTravelled, "totalT": totalTimeTaken, "maxV": maxV, "cutOffT": cutOffTime, "aveA": aveA, "maxA": maxA}
    
    return data


# # 5) Generating and Saving results

# In[ ]:


def saveDataInsights(data, title, output_path):
    '''
    Input: Dictionary of data generated from data(df)
    Output: A .txt file located under specified_output_path/{title_of_csv_file}-insights with all insights generated
    '''
    
    result_data = f'''
    # Total distance travelled: {data["d"]} m
    # Total time taken (For multiple readings, it is the longest time): {data["totalT"]} s
    # Max velocity: {data["maxV"]} m/s
    # Cut-off time for acceleration: {data["cutOffT"]} s
    # Average acceleration: {data["aveA"]} m/s^2
    # Max acceleration: {data["maxA"]} m/s^2
    '''
    
    # Save Data Insights to File
    formatted_output_path = output_path + "/" + title + "-insights.txt"
    insights_file = open(formatted_output_path,"w",newline='')
    insights_file.write(result_data)
    insights_file.close()


# In[ ]:


def plotAndSaveGraphs(df, title, data, output_path):
    # Plot A/t
    args_at = {"title": title, "type": "Acceleration", "xlabel": "time (s)", "ylabel": "Y-Axis Acceleration (m/s^2)", "xcoldata": "TimeElapsed", "ycoldata": "Acc-Y-Adjusted", "indicatorName": "Acceleration Cut-off", "indicatorData": data["cutOffT"], "path": output_path}
    plotAndSave(df, args_at)
    
    # Plot & Save V/t
    args_vt = {"title": title, "type": "Velocity", "xlabel": "time (s)", "ylabel": "Velocity (m/s)", "xcoldata": "TimeElapsed", "ycoldata": "V(t)", "indicatorName": "Max Velocity", "indicatorData": data["maxV"], "path": output_path}
    plotAndSave(df, args_vt)
    
    # Plot & Save S/t
    args_st = {"title": title, "type": "Displacement", "xlabel": "time (s)", "ylabel": "Displacement (m)", "xcoldata": "TimeElapsed", "ycoldata": "S(t)", "indicatorName": "Total Distance Travelled", "indicatorData": data["d"], "path": output_path}
    plotAndSave(df, args_st)


# # 6) Prompt user for input

# In[ ]:


def userInput(): 
     # i.e. "Output-1"
    output_dir = "/" + input("Please input the relative path ending with the desired folder name for output: ") 
    user_dir = input("Please input the relative path to the directory containing a single set of .csv files: ")
    while True:
        try:                
            smoothen_curve = input("Enter the step value for smoothing [1 - 15]\n\t The higher the value, the smoother it is: ")
            if int(smoothen_curve) in range(1,16):
                pass
            else:
                continue
            format_acc = input("Is your acceleration measured in g? (I.e. '1' represents 9.81ms^-2) [Y]/[N]: ").lower()

            if format_acc == 'y' or format_acc == 'n':
                pass
            else:
                continue
            is_ms = input("Is your time measured in ms? (I.e. '1000' represents 1s) [Y]/[N]: ").lower()
            if is_ms == 'y' or is_ms == 'n':
                pass
            else:
                continue
            threshold = float(input("Enter the threshold value for automatic start point truncation [0 - 1]\n\t Typically 0.02: "))
            break
        except:
            print("Invalid Input. Please try again.")
            continue
            
    return output_dir, user_dir, smoothen_curve, format_acc, is_ms, threshold


# ## Run it all (Main function)

# In[ ]:


def main():
    # Get user input
    output_dir, user_dir, smoothen_curve, format_acc, is_ms, threshold = userInput()

    # Setup Output Path
    my_path = os.path.realpath("")
    output_path = my_path + output_dir
    mkdir_p(output_path)
    
    # Create new Excel workbook
    wb = openpyxl.Workbook()
    # Just to remove the spare sheet
    wb.remove(wb.active)
    
    

    # Retrieve directory containing all .csv files
    path = "c:\\"
    extension = "csv"
    os.chdir(user_dir)
    all_files = glob.glob(f"*.{extension}")
    d, file_names = sortedDict(all_files)
      
    print(f"Results will be available at: {output_path}")


    # Reads sorted dict of csv files. Converts them to data frame and concatenate similar ones together. Outputs excel wb
    excel_wb_name = csvToExcel(d, output_path, is_ms)
    
    # Change back to original directory
    os.chdir(my_path)
    
    acceleration_data = pd.read_excel(f"{excel_wb_name}.xlsx", sheet_name=None, header=None)

    for i, title in enumerate(acceleration_data):
        print(f"Status: Running {i+1} / {len(acceleration_data)}")
        df = acceleration_data[title]
        
        isInG = (format_acc == "y")
        dfFormatArgs = {"Threshold": threshold, "isInG": isInG}
        # Calculations to get the adjusted acceleration, velocity, and displacement
        df = dfFormat(df, smoothen_curve, format_acc, dfFormatArgs) 

        # Save the DataFrame as the csv file
        formatted_csv_path = output_path + "/" + title + "-formatted.csv"
        df.to_csv(formatted_csv_path)
        
        # Get data insights
        data = getDataInsights(df)
        
        # Save the insights into a .txt file
        saveDataInsights(data, title, output_path)
        
        # Plot & save data to specified output location
        plotAndSaveGraphs(df, title, data, output_path)

        print("Program has finished running!")
        print("Results will be available at: {}".format(output_path))


# In[ ]:


main()

